# import openpyxl
import openpyxl


class LoginPageData:

    test_LoginPage_invalid_data = [
        {"username": "amin", "password": "13", "ExpectedResult": "No user have been register into the system"}
    ]

    @staticmethod
    def getValidData(PATH):
        Dict = []
        book = openpyxl.load_workbook(PATH)
        sheet = book.active

        data = {}
        for j in range(3, sheet.max_row + 1):
            if sheet.cell(row=j, column=5).value is not None:
                data[sheet.cell(row=j, column=5).value] = str(sheet.cell(row=j, column=6).value)
        Dict.append(data)

        return Dict

