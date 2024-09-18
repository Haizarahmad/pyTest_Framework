import openpyxl


class AhliKariahPageData:

    @staticmethod
    def getData(PATH):
        Dict = []
        book = openpyxl.load_workbook(PATH)
        sheet = book.active

        data = {}
        for j in range(3, sheet.max_row + 1):
            if sheet.cell(row=j, column=5).value is not None:
                data[sheet.cell(row=j, column=5).value] = str(sheet.cell(row=j, column=6).value)
        Dict.append(data)

        return Dict

    @staticmethod
    def getDataAmend(PATH):
        #("num1, num2, expected_total,
        #[
        # ("10", "20", "12"),
        # ("13", "12", "23")
        #])
        fields = ["num1", "num2", "num3"]
        Dict = []
        book = openpyxl.load_workbook(PATH)
        sheet = book.active

        data = {}
        for j in range(3, sheet.max_row + 1):
            if sheet.cell(row=j, column=5).value is not None:
                data[sheet.cell(row=j, column=5).value] = str(sheet.cell(row=j, column=6).value)
        Dict.append(data)

        return Dict

